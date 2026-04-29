# -*- coding: utf-8 -*-
import openpyxl

def create_full_admin_tc():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    P = 'nowSOS - Hệ thống Điều phối Cứu hộ Thông minh'
    
    def sh(wb, name, mod, rows):
        ws = wb.create_sheet(name)
        ws['A1'] = 'Project Name'; ws['B1'] = P
        ws['A2'] = 'Module'; ws['B2'] = mod
        
        ws['B3']='Passed'; ws['C3']='Failed'; ws['D3']='Untested'; ws['E3']='Blocked'; ws['F3']='Total'
        for r, lb in [(4,'Round 1'), (5,'Round 2')]:
            c = 'G' if r==4 else 'J'
            ws[f'A{r}']=lb; ws[f'B{r}']=f'=COUNTIF({c}16:{c}100,"Passed")'
            ws[f'C{r}']=f'=COUNTIF({c}16:{c}100,"Failed")'; ws[f'D{r}']=f'=COUNTIF({c}16:{c}100,"Untested")'
            ws[f'E{r}']=f'=COUNTIF({c}16:{c}100,"Blocked")'; ws[f'F{r}']=f'=COUNTA(A16:A100)'
            
        ws['A12']='Test Case ID'; ws['B12']='Mô tả'; ws['C12']='Các bước thực hiện'
        ws['D12']='Điều kiện'; ws['E12']='Kết quả mong đợi'; ws['F12']='Kết quả thực tế'
        ws['G12']='Kết quả Vòng 1'; ws['J12']='Kết quả Vòng 2'
        
        ws['G13']='Trạng thái'; ws['H13']='Ngày'; ws['I13']='Người test'
        ws['J13']='Trạng thái'; ws['K13']='Ngày'; ws['L13']='Người test'
        
        r = 15
        for row in rows:
            if len(row) == 1:
                ws.cell(r, 1).value = row[0]
            else:
                for c2, v in enumerate(row, 1):
                    ws.cell(r, c2).value = v
                ws.cell(r, 7).value = 'Untested'
            r += 1

    # 1. Dashboard
    sh(wb, 'Dashboard', 'Dashboard tổng quan', [
        ['KIỂM THỬ GIAO DIỆN & NÚT BẤM'],
        ('TC-DB01', 'Kiểm tra hiển thị bố cục Dashboard', '1. Đăng nhập Admin\n2. Xem trang Dashboard', 'Admin đã đăng nhập', 'Hiển thị đủ 4 widget thống kê, 2 biểu đồ và 2 bảng dữ liệu'),
        ('TC-DB02', 'Kiểm tra hoạt động nút "Làm mới"', '1. Click nút "Làm mới" ở góc phải', 'Đang ở Dashboard', 'Icon quay tròn, dữ liệu các ô widget được tải lại chính xác'),
        ('TC-DB03', 'Kiểm tra click nút "Điều phối" trên bảng YC chờ', '1. Click nút "Điều phối" tại một dòng', 'Có YC đang chờ', 'Chuyển hướng sang trang assignments với ID tương ứng'),
        ('TC-DB04', 'Kiểm tra click "Xem tất cả" hàng đợi', '1. Click link "Xem tất cả"', 'Đang ở Dashboard', 'Chuyển hướng sang trang Hàng Đợi theo ưu tiên'),
        ('TC-DB05', 'Kiểm tra bộ lọc "Hôm nay"', '1. Click nút "Hôm nay"', 'Mặc định', 'Biểu đồ sự cố chỉ hiển thị dữ liệu trong ngày'),
        ('TC-DB06', 'Kiểm tra bộ lọc "7 ngày"', '1. Click nút "7 ngày"', 'Đang ở Dashboard', 'Biểu đồ cập nhật số liệu 7 ngày gần nhất'),
    ])

    # 2. Hàng đợi
    sh(wb, 'HangDoi', 'Hàng đợi theo ưu tiên', [
        ['KIỂM THỬ CHỨC NĂNG HÀNG ĐỢI'],
        ('TC-Q01', 'Kiểm tra nút "Đồng bộ" danh sách', '1. Click "Đồng bộ" ở góc phải', 'Đang ở trang Hàng đợi', 'Dữ liệu YC mới từ người dân được nạp lại'),
        ('TC-Q02', 'Kiểm tra nút "Phân Bổ Ngay" trên card YC', '1. Click nút "Phân Bổ Ngay" dưới card', 'Có YC trong hàng đợi', 'Chuyển sang màn hình điều phối cho YC đó'),
        ('TC-Q03', 'Kiểm tra hiển thị badge mức độ khẩn cấp', '1. Quan sát góc trái card YC', 'Có YC mức HIGH/CRITICAL', 'Màu sắc badge hiển thị đúng: đỏ đậm cho Critical, đỏ tươi cho High'),
    ])

    # 3. Theo dõi cứu hộ
    sh(wb, 'TheoDoi', 'Theo dõi cứu hộ (LIVE)', [
        ['KIỂM THỬ TRANG THEO DÕI LIVE'],
        ('TC-TD01', 'Kiểm tra hiển thị danh sách YC ĐANG THEO DÕI', '1. Vào trang "Theo dõi cứu hộ"\n2. Xem danh sách bên trái', 'Hệ thống có YC đang xử lý', 'Hiển thị badge số lượng (ví dụ: 9) và các card YC'),
        ('TC-TD02', 'Kiểm tra click chọn YC để xem chi tiết', '1. Click chọn YC #30', 'Danh sách có nhiều YC', 'Ô bên phải cập nhật đúng: Lũ lụt, Hoàng Thị Hạnh, 0905555555'),
        ('TC-TD03', 'Kiểm tra nút "Đồng bộ" live data', '1. Click nút "Đồng bộ" trên cùng', 'Đang ở trang', 'Tải lại danh sách live không bị giật lag'),
        ('TC-TD04', 'Kiểm tra nút "X" đóng panel đơn vị cứu hộ', '1. Click nút "X" tại panel "THÔNG TIN ĐƠN VỊ CỨU HỘ"', 'Đã chọn 1 YC', 'Ẩn/Đóng thông tin đơn vị cứu hộ tạm thời'),
    ])

    # 4. Phân công
    sh(wb, 'PhanCong', 'Phân công đội cứu hộ', [
        ['KIỂM THỬ ĐIỀU PHỐI ĐỘI'],
        ('TC-PC01', 'Kiểm tra nút "Chọn tất cả" đội', '1. Click "Chọn tất cả" ở góc phải', 'Đã chọn 1 YC xử lý', 'Tất cả các đội cứu hộ đủ điều kiện được tick chọn'),
        ('TC-PC02', 'Kiểm tra nút "Bỏ chọn" đội', '1. Click "Bỏ chọn"', 'Đang chọn nhiều đội', 'Xóa toàn bộ các đội đang chọn về trạng thái trống'),
        ('TC-PC03', 'Kiểm tra click "Chốt & Xuất Phát Lệnh"', '1. Chọn 1 đội\n2. Click "Chốt & Xuất Phát Lệnh"', 'Đã chọn YC và đội', 'Nút hiện spinner, gửi API thành công và chuyển trạng thái nhiệm vụ'),
    ])

    wb.save(r'C:\xampp\htdocs\DATN\GR29\TC_Admin_Full_SatCode.xlsx')

if __name__ == '__main__':
    create_full_admin_tc()

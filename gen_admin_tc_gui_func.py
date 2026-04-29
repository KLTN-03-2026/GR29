# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_pro_admin_tc():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    cat_font = Font(bold=True, color="1E3A8A", size=12)
    cat_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def sh(wb, name, title, test_cases):
        ws = wb.create_sheet(name)
        
        # Header Info
        ws['A1'] = 'DỰ ÁN'
        ws['B1'] = 'nowSOS - Hệ thống Điều phối Cứu hộ Thông minh'
        ws['A2'] = 'PHÂN HỆ'
        ws['B2'] = f'ADMIN CONSOLE - {title}'
        
        for r in [1, 2]:
            ws[f'A{r}'].font = Font(bold=True)
        
        # Test Case Table Headers
        headers = ['ID', 'Loại Test', 'Chức năng / Nút bấm', 'Mô tả kịch bản (Steps)', 'Điều kiện', 'Kết quả mong đợi', 'Trạng thái', 'Ghi chú']
        col_widths = [10, 15, 25, 45, 20, 40, 15, 20]
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = h
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = col_widths[col-1]
            
        r = 5
        current_type = None
        for tc in test_cases:
            if len(tc) == 1:
                # Category row
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
                cell = ws.cell(row=r, column=1)
                cell.value = tc[0]
                cell.font = cat_font
                cell.fill = cat_fill
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border
                current_type = tc[0]
                r += 1
            else:
                # tc format: (ID, Target, Steps, Condition, Expected)
                t_type = "GUI_SHOW" if "GUI" in current_type else "FUNCTION_SHOW"
                row_data = [tc[0], t_type, tc[1], tc[2], tc[3], tc[4], 'Untested', '']
                for col, val in enumerate(row_data, 1):
                    cell = ws.cell(row=r, column=col)
                    cell.value = val
                    cell.alignment = align_left if col in [3,4,5,6] else align_center
                    cell.border = border
                r += 1

    # 1. Dashboard
    sh(wb, '1.Dashboard', 'Dashboard tổng quan', [
        ['[1] - KIỂM THỬ GIAO DIỆN (GUI_SHOW)'],
        ('TC_DB_01', '4 Widget thống kê', 'Kiểm tra hiển thị 4 thẻ: Chờ xử lý, Đang xử lý, Hoàn thành, Đội sẵn sàng.', 'Đăng nhập Admin', 'Hiển thị đủ 4 thẻ kèm icon tương ứng, số liệu không bị tràn khung.'),
        ('TC_DB_02', 'Biểu đồ phân bổ sự cố', 'Kiểm tra phần render của biểu đồ "Phân bố sự cố theo loại".', 'Có dữ liệu', 'Biểu đồ tròn/donut hiển thị rõ ràng, màu sắc phân biệt các loại sự cố.'),
        ('TC_DB_03', 'Bảng Yêu cầu chờ xử lý', 'Kiểm tra lưới danh sách 5 yêu cầu mới nhất ở cuối trang.', 'Có yêu cầu chờ', 'Có các cột: Mã, Loại, Vị trí, Mức độ, Thời gian và nút "Điều phối".'),
        ['[2] - KIỂM THỬ CHỨC NĂNG (FUNCTION_SHOW)'],
        ('TC_DB_04', 'Nút "Làm mới"', 'Click vào nút Làm mới ở góc trên bên phải.', 'Đang ở Dashboard', 'Nút hiển thị spinner (spin css), dữ liệu các widget và biểu đồ được gọi lại API.'),
        ('TC_DB_05', 'Bộ lọc thời gian', 'Lần lượt click vào các nút: Hôm nay, 7 ngày, 30 ngày.', 'Đang ở Dashboard', 'Dữ liệu biểu đồ thay đổi tương ứng với khoảng thời gian đã chọn.'),
        ('TC_DB_06', 'Nút "Xem tất cả"', 'Click link "Xem tất cả" ở bảng Yêu cầu chờ xử lý.', 'Có danh sách', 'Hệ thống chuyển hướng (router.push) sang trang Hàng đợi ưu tiên.'),
        ('TC_DB_07', 'Nút "Điều phối"', 'Click nút "Điều phối" trên một dòng cụ thể.', 'Dòng có ID cụ thể', 'Chuyển sang trang Phân công với ID sự cố tương ứng.'),
    ])

    # 2. Hàng đợi (Queue)
    sh(wb, '2.HangDoi', 'Hàng đợi theo ưu tiên', [
        ['[1] - KIỂM THỬ GIAO DIỆN (GUI_SHOW)'],
        ('TC_HD_01', 'Empty State', 'Kiểm tra hiển thị khi không có yêu cầu nào.', 'Hàng đợi rỗng', 'Hiển thị icon ly cafe (hoặc tương tự) và dòng "Không có yêu cầu cần phân bổ".'),
        ('TC_HD_02', 'Danh sách Card Yêu cầu', 'Quan sát các thẻ Yêu cầu trên màn hình.', 'Có YC', 'Mỗi YC hiển thị thành 1 card, có Ảnh sự cố (nếu có), Badge ưu tiên (Đỏ/Vàng/Xanh).'),
        ('TC_HD_03', 'Màu sắc Mức độ', 'Kiểm tra màu sắc của badge Mức độ ưu tiên.', 'Có đủ 4 mức', 'CRITICAL: Đỏ đậm; HIGH: Đỏ; MEDIUM: Vàng; LOW: Xanh lá.'),
        ['[2] - KIỂM THỬ CHỨC NĂNG (FUNCTION_SHOW)'],
        ('TC_HD_04', 'Nút "Đồng bộ"', 'Click nút "Đồng bộ" ở góc trên.', 'Bất kỳ', 'Gọi lại API lấy danh sách mới nhất, icon xoay (spin).'),
        ('TC_HD_05', 'Nút "Phân Bổ Ngay"', 'Click nút "Phân Bổ Ngay" tại 1 card yêu cầu.', 'Trên card YC', 'Chuyển hướng (push) sang trang Assignments để tiến hành điều phối.'),
    ])

    # 3. Đang Xử Lý (DangXuLy)
    sh(wb, '3.DangXuLy', 'Đang Xử Lý', [
        ['[1] - KIỂM THỬ GIAO DIỆN (GUI_SHOW)'],
        ('TC_DXL_01', 'UI Bộ lọc Tìm kiếm', 'Kiểm tra vùng tìm kiếm chung, Loại sự cố, Thời gian.', 'Truy cập trang', 'Hiển thị 3 input lọc và 1 nút "Xóa lọc" ở thanh ngang trên cùng.'),
        ('TC_DXL_02', 'Hiển thị Card Đang xử lý', 'Kiểm tra thẻ nhiệm vụ đang thực hiện.', 'Có YC đang xử lý', 'Hiển thị ảnh (hoặc placeholder text nếu không có ảnh), tiêu đề, địa chỉ (truncate 2 dòng).'),
        ['[2] - KIỂM THỬ CHỨC NĂNG (FUNCTION_SHOW)'],
        ('TC_DXL_03', 'Lọc theo text', 'Nhập mã sự cố hoặc địa chỉ vào ô "Tìm kiếm chung".', 'Nhập liệu', 'Danh sách lọc realtime ngay khi gõ (input event) mà không cần nhấn Enter.'),
        ('TC_DXL_04', 'Lọc Loại sự cố', 'Chọn 1 loại sự cố cụ thể trong Dropdown.', 'Chọn Dropdown', 'Danh sách chỉ hiển thị các yêu cầu thuộc loại tương ứng.'),
        ('TC_DXL_05', 'Nút "Xóa lọc"', 'Nhập giá trị vào bộ lọc rồi nhấn "Xóa lọc".', 'Đã lọc', 'Khôi phục toàn bộ danh sách, xóa trắng các ô input/dropdown.'),
        ('TC_DXL_06', 'Nút "Theo dõi nhiệm vụ"', 'Nhấn nút "Theo dõi nhiệm vụ" trên thẻ.', 'Click button', 'Chuyển hướng sang trang Theo dõi cứu hộ truyền kèm ID.'),
    ])

    # 4. Theo Dõi (TheoDoiYeuCau)
    sh(wb, '4.TheoDoi', 'Theo dõi cứu hộ', [
        ['[1] - KIỂM THỬ GIAO DIỆN (GUI_SHOW)'],
        ('TC_TD_01', 'Cột trái: Danh sách theo dõi', 'Kiểm tra danh sách "Yêu cầu đang theo dõi".', 'Có YC', 'Hiển thị dạng list dọc, có badge LIVE và badge báo số lượng.'),
        ('TC_TD_02', 'Cột phải: Empty State', 'Kiểm tra khi chưa chọn yêu cầu nào ở cột trái.', 'Chưa select', 'Hiển thị thông báo "Chọn yêu cầu để theo dõi" ở khung phải.'),
        ('TC_TD_03', 'Cột phải: Timeline', 'Kiểm tra khu vực Timeline tiến độ.', 'Đã chọn YC', 'Hiển thị các bước: Tiếp nhận, Phân công, Di chuyển, Tới hiện trường, Hoàn thành. Bước hiện tại nháy đỏ.'),
        ['[2] - KIỂM THỬ CHỨC NĂNG (FUNCTION_SHOW)'],
        ('TC_TD_04', 'Chọn 1 Yêu cầu', 'Click vào dòng yêu cầu `#30` ở cột trái.', 'Click list', 'Khung bên phải load dữ liệu chi tiết của `#30` (Tên đội, KH, bản đồ). Dòng bên trái sáng lên (active).'),
        ('TC_TD_05', 'Nút "X" đóng panel Đội', 'Click dấu "X" ở góc phải phần "Thông tin đơn vị cứu hộ".', 'Đang xem YC', 'Hành động clearSelection() được kích hoạt, khung phải quay về Empty state.'),
        ('TC_TD_06', 'Nút "Liên hệ khách hàng"', 'Click nút gọi điện thoại.', 'Trên Mobile/Desktop có app', 'Mở ứng dụng gọi điện mặc định với số của người gửi yêu cầu.'),
        ('TC_TD_07', 'Nút "Huỷ nhiệm vụ"', 'Click "Huỷ nhiệm vụ".', 'Click button', 'Hiển thị confirm alert, nếu Yes -> gọi API Hủy (hiện cảnh báo nếu chưa code xong).'),
    ])

    # 5. Phân Công (Assignments)
    sh(wb, '5.PhanCong', 'Phân công đội cứu hộ', [
        ['[1] - KIỂM THỬ GIAO DIỆN (GUI_SHOW)'],
        ('TC_PC_01', 'Thông tin sự cố', 'Kiểm tra box hiển thị thông tin sự cố cần phân công.', 'Truy cập trang', 'Hiển thị rõ mã YC, mức độ, tọa độ, mô tả sự cố để người điều phối nắm tình hình.'),
        ('TC_PC_02', 'Danh sách gợi ý đội', 'Kiểm tra bảng danh sách các đội cứu hộ.', 'Có đội sẵn sàng', 'Bảng hiển thị Tên đội, Loại xử lý được, Khoảng cách (km), Thanh Capacity (Công suất).'),
        ('TC_PC_03', 'Thanh Capacity (Công suất)', 'Kiểm tra thanh quá tải của đội cứu hộ.', 'Đội đang làm', 'Nếu đang xử lý >3 vụ, thanh Capacity chuyển đỏ. Trạng thái hiện "Overload".'),
        ['[2] - KIỂM THỬ CHỨC NĂNG (FUNCTION_SHOW)'],
        ('TC_PC_04', 'Giải thuật gợi ý (Sắp xếp)', 'Quan sát thứ tự đội.', 'Tự động', 'Đội cùng loại + cùng quận xếp đầu, tiếp đến cùng loại, và xếp theo khoảng cách từ thấp đến cao.'),
        ('TC_PC_05', 'Click chọn Đội cứu hộ', 'Tick vào checkbox của 1 đội không quá tải.', 'Đội hợp lệ', 'Checkbox được đánh dấu, Đội được add vào mảng selectedTeams.'),
        ('TC_PC_06', 'Click chọn Đội quá tải', 'Cố tình tick vào đội báo Overload.', 'Đội quá tải', 'Hệ thống disable dòng, không cho phép tick chọn.'),
        ('TC_PC_07', 'Nút "Chọn tất cả" / "Bỏ chọn"', 'Sử dụng 2 nút tiện ích.', 'Trên thanh công cụ', 'Tick/Untick hàng loạt toàn bộ đội cứu hộ đang hiển thị trên grid.'),
        ('TC_PC_08', 'Nút "Chốt & Xuất Phát Lệnh"', 'Chọn 1 YC, chọn 1 Đội, click Chốt.', 'Đầy đủ đk', 'Nút disable và hiện Spinner. Gọi API. Hiện Toast Success. Chuyển hướng về bảng Đang xử lý.'),
    ])

    # 6. Đã Hoàn Thành (DaHoanThanh)
    sh(wb, '6.DaHoanThanh', 'Đã Hoàn Thành', [
        ['[1] - KIỂM THỬ GIAO DIỆN (GUI_SHOW)'],
        ('TC_HT_01', 'Bảng lịch sử', 'Kiểm tra lưới lịch sử cứu hộ.', 'Trang Đã hoàn thành', 'Hiển thị danh sách dạng bảng, có thời gian hoàn thành cụ thể.'),
        ['[2] - KIỂM THỬ CHỨC NĂNG (FUNCTION_SHOW)'],
        ('TC_HT_02', 'Tìm kiếm Lịch sử', 'Nhập text vào thanh tìm kiếm.', 'Nhập text', 'Lọc danh sách các yêu cầu đã đóng theo từ khóa.'),
    ])

    # 7. Heatmap & 8. Tracking
    sh(wb, '7_8.Map&Tracking', 'Bản đồ Nhiệt & Vị trí', [
        ['[1] - KIỂM THỬ GIAO DIỆN (GUI_SHOW)'],
        ('TC_MAP_01', 'Render Heatmap', 'Truy cập "Bản đồ nhiệt nguy hiểm".', 'Load trang', 'Bản đồ tải thành công, vùng nhiều sự cố có vệt màu Đỏ/Cam (Heat).'),
        ('TC_MAP_02', 'Render Marker Tracking', 'Truy cập "Vị trí đội".', 'Load trang', 'Hiển thị các icon xe cứu hộ rải rác trên bản đồ theo tọa độ GPS.'),
        ['[2] - KIỂM THỬ CHỨC NĂNG (FUNCTION_SHOW)'],
        ('TC_MAP_03', 'Click Marker', 'Click vào 1 icon xe cứu hộ.', 'Bản đồ', 'Bật lên Popup (Tooltip) hiển thị Tên đội, Trạng thái, Vị trí đang xử lý.'),
    ])

    # 9. Báo Cáo
    sh(wb, '9.BaoCao', 'Báo cáo thống kê', [
        ['[1] - KIỂM THỬ GIAO DIỆN (GUI_SHOW)'],
        ('TC_BC_01', 'UI Thống kê', 'Kiểm tra các biểu đồ thống kê.', 'Trang Báo Cáo', 'Hiển thị biểu đồ cột, biểu đồ tròn tổng kết.'),
        ['[2] - KIỂM THỬ CHỨC NĂNG (FUNCTION_SHOW)'],
        ('TC_BC_02', 'Xuất báo cáo', 'Click nút xuất PDF / Excel.', 'Nếu có nút', 'Trình duyệt tải xuống file với tên chứa ngày tháng hiện tại.'),
    ])

    # 10. Cấu hình & Tài khoản
    sh(wb, '10.CauHinh', 'Cấu hình và Tài khoản', [
        ['[1] - KIỂM THỬ GIAO DIỆN (GUI_SHOW)'],
        ('TC_CH_01', 'Giao diện Danh mục', 'Truy cập menu Loại sự cố / Đội cứu hộ.', 'Quyền Admin', 'Hiển thị bảng dữ liệu (Read-only mode nếu đang khóa chức năng sửa).'),
        ('TC_CH_02', 'Giao diện Trọng số AI', 'Truy cập "Trọng số AI scoring".', 'Quyền Admin', 'Hiển thị các thanh trượt (slider) cho: Cùng loại, Khoảng cách, Hiệu suất.'),
        ('TC_CH_03', 'Giao diện Tài khoản', 'Truy cập quản lý Admin/User/Rescuer.', 'Quyền Admin', 'Hiển thị lưới dữ liệu danh sách user.'),
        ['[2] - KIỂM THỬ CHỨC NĂNG (FUNCTION_SHOW)'],
        ('TC_CH_04', 'Chỉnh sửa Trọng số AI', 'Kéo thanh trượt khoảng cách lên 80%, ấn Lưu.', 'Đang ở Trọng số', 'Hệ thống thông báo "Cập nhật thành công", thuật toán sẽ đổi độ ưu tiên.'),
        ('TC_CH_05', 'Thao tác Tài khoản', 'Test nút Khóa/Mở tài khoản nếu có.', 'Trang User', 'Hệ thống đổi trạng thái và không cho user đăng nhập nữa.'),
    ])

    wb.save(r'C:\xampp\htdocs\DATN\GR29\ProjectTestCase_Admin_Pro.xlsx')
    print("Done generating PRO Excel Test Cases.")

if __name__ == '__main__':
    create_pro_admin_tc()


import openpyxl

PROJECT = 'Thiet ke va phat trien he thong dieu phoi cuu ho thong minh trong thien tai dua tren phan tich muc do khan cap'

def make_header(ws, module_code):
    ws['A1'] = 'Project Name'; ws['B1'] = PROJECT
    ws['A2'] = 'Module Code'; ws['B2'] = module_code
    ws['B3'] = 'Hoan thanh'; ws['C3'] = 'Loi'; ws['D3'] = 'Chua kiem tra'
    ws['E3'] = 'Bi chan'; ws['F3'] = 'Tong so truong hop thi nghiem'
    ws['A4'] = 'Round 1'
    ws['B4'] = '=COUNTIF(G16:G60,"Passed")'; ws['C4'] = '=COUNTIF(G16:G60,"Failed")'
    ws['D4'] = '=COUNTIF(G16:G60,"Untested")'; ws['E4'] = '=COUNTIF(G16:G60,"Blocked")'
    ws['F4'] = '=COUNTA(A16:A60)'
    ws['A5'] = 'Round 2'
    ws['B5'] = '=COUNTIF(J16:J60,"Passed")'; ws['C5'] = '=COUNTIF(J16:J60,"Failed")'
    ws['D5'] = '=COUNTIF(J16:J60,"Untested")'; ws['E5'] = '=COUNTIF(J16:J60,"Blocked")'
    ws['F5'] = '=COUNTA(A16:A60)'
    ws['A12'] = 'Test Case ID'; ws['B12'] = 'Mo ta'; ws['C12'] = 'Hanh dong'
    ws['D12'] = 'Dieu kien tien quyet'; ws['E12'] = 'Ket qua mong doi'
    ws['F12'] = 'Ket qua thuc te'; ws['G12'] = 'Ket qua'; ws['J12'] = 'Ket qua'; ws['M12'] = 'Chu thich'
    ws['G13'] = 'Vong 1'; ws['J13'] = 'Vong 2'
    ws['G14'] = 'Trang thai'; ws['H14'] = 'Ngay kiem tra'; ws['I14'] = 'Nguoi kiem tra'
    ws['J14'] = 'Trang thai'; ws['K14'] = 'Ngay kiem tra'; ws['L14'] = 'Nguoi kiem tra'

def write_rows(ws, rows):
    r = 15
    for row in rows:
        if len(row) == 1:
            ws.cell(row=r, column=1).value = row[0]
        else:
            cols = ['tc_id','mo_ta','hanh_dong','dieu_kien','mong_doi','thuc_te','trang_thai']
            data = list(row) + [''] * (7 - len(row))
            ws.cell(row=r, column=1).value = data[0]
            ws.cell(row=r, column=2).value = data[1]
            ws.cell(row=r, column=3).value = data[2]
            ws.cell(row=r, column=4).value = data[3]
            ws.cell(row=r, column=5).value = data[4]
            ws.cell(row=r, column=6).value = data[5] if len(data) > 5 else ''
            ws.cell(row=r, column=7).value = data[6] if len(data) > 6 else 'Untested'
        r += 1

wb = openpyxl.load_workbook(r'C:\xampp\htdocs\DATN\GR29\8.1.ProjectTestCaseSprint1(CHUAXONG).xlsx')

# ========== SHEET: DANG XUAT ==========
ws = wb.create_sheet('Dang Xuat')
make_header(ws, 'Dang Xuat')
write_rows(ws, [
    ['GUI_SHOW Trang dang xuat / Menu'],
    ('GUI-DX01','[Menu nguoi dung] Nut Dang Xuat','Kiem tra UI menu','Da dang nhap he thong','-Hien thi tuy chon "Dang Xuat" trong menu/avatar\n-Status: enable','','Untested'),
    ('GUI-DX02','[Hop thoai xac nhan] Dialog','Kiem tra dialog xac nhan','Click nut Dang Xuat','-Title: "Xac nhan dang xuat"\n-Co nut "Xac nhan" va "Huy"\n-Khong tu dong dong\n-Status: enable','','Untested'),
    ('GUI-DX03','[Icon dang xuat] Trong sidebar/topbar','Kiem tra hien thi icon','Da dang nhap','-Icon dang xuat hien thi ro rang\n-Co tooltip khi hover','','Untested'),
    ['FUNCTION_SHOW Chuc nang dang xuat'],
    ('FUNC-DX01','Dang xuat thanh cong','1. Click nut "Dang Xuat"\n2. Click "Xac nhan" trong dialog','Da dang nhap he thong','Dang xuat thanh cong, session bi xoa, chuyen ve man hinh Dang Nhap','','Untested'),
    ('FUNC-DX02','Huy dang xuat - chon "Huy"','1. Click nut "Dang Xuat"\n2. Click "Huy" trong dialog','Da dang nhap he thong','He thong giu nguyen trang thai dang nhap, dialog dong lai','','Untested'),
    ('FUNC-DX03','Huy dang xuat - bam phim Esc','1. Click nut "Dang Xuat"\n2. Bam phim Esc','Da dang nhap he thong','Dialog dong lai, he thong giu nguyen trang thai dang nhap','','Untested'),
    ('FUNC-DX04','Sau dang xuat khong the truy cap trang can dang nhap','1. Dang xuat thanh cong\n2. Nhap truc tiep URL trang chinh vao trinh duyet','Da dang xuat','He thong tu dong chuyen huong ve trang Dang Nhap','','Untested'),
    ('FUNC-DX05','Token / Session bi huy sau dang xuat','1. Dang xuat\n2. Dung DevTools / Postman goi API voi token cu','Da dang xuat','API tra ve 401 Unauthorized, token khong con hieu luc','','Untested'),
    ('FUNC-DX06','Dang xuat tren mot tab, kiem tra tab khac','1. Mo 2 tab cung dang nhap\n2. Dang xuat tren Tab 1\n3. Thao tac tren Tab 2','Da dang nhap tren 2 tab','Tab 2 chuyen ve man hinh dang nhap hoac thong bao het phien','','Untested'),
    ('FUNC-DX07','Dang xuat khi mat mang','1. Tat ket noi mang\n2. Click Dang Xuat\n3. Xac nhan','Da dang nhap, mat ket noi','He thong xoa session cuc bo, chuyen ve man hinh Dang Nhap khi co mang','','Untested'),
])

# ========== SHEET: GUI YEU CAU CUU HO ==========
ws2 = wb.create_sheet('Gui Yeu Cau Cuu Ho')
make_header(ws2, 'Gui Yeu Cau Cuu Ho')
write_rows(ws2, [
    ['GUI_SHOW Trang gui yeu cau cuu ho'],
    ('GUI-GYC01','[Title] Tieu de trang gui yeu cau','Kiem tra giao dien tieu de','Da dang nhap, la nguoi dan','-Title: "Gui yeu cau cuu ho"\n-Hien thi ro rang tren dau trang\n-Status: enable','','Untested'),
    ('GUI-GYC02','[Loai su co] Dropdown chon loai su co','Kiem tra dropdown','Da dang nhap','-Dropdown hien thi cac loai: Lu lut, Chay, Dong dat, Sat lo, Bao...\n-Co gia tri mac dinh "Chon loai su co"\n-Status: enable','','Untested'),
    ('GUI-GYC03','[Mo ta tinh huong] Textbox nhap mo ta','Kiem tra o mo ta','Da dang nhap','-Placeholder: "Mo ta chi tiet tinh huong khan cap"\n-Cho phep nhap nhieu dong\n-Gioi han 500 ky tu\n-Status: enable','','Untested'),
    ('GUI-GYC04','[Vi tri] Field vi tri + nut lay GPS','Kiem tra vi tri field','Da dang nhap','-Hien thi vi tri GPS hien tai hoac cho phep nhap thu cong\n-Nut "Lay vi tri hien tai"\n-Hien thi toa do hoac dia chi\n-Status: enable','','Untested'),
    ('GUI-GYC05','[So nguoi can ho tro] Input so nguoi','Kiem tra input so luong','Da dang nhap','-Placeholder: "Nhap so nguoi can ho tro"\n-Chi chap nhan so nguyen duong\n-Status: enable','','Untested'),
    ('GUI-GYC06','[Muc do khan cap] Radio / Dropdown','Kiem tra chon muc do','Da dang nhap','-Cac muc: Khan cap, Cao, Trung binh, Thap\n-Mac dinh: Trung binh\n-Status: enable','','Untested'),
    ('GUI-GYC07','[Anh/Video minh chung] Upload component','Kiem tra upload','Da dang nhap','-Nut "Chon anh / video"\n-Hien thi preview anh da chon\n-Cho phep chon nhieu anh (toi da 5)\n-Status: enable','','Untested'),
    ('GUI-GYC08','[Gui yeu cau] Button','Kiem tra nut gui','Da dang nhap','-Text: "Gui yeu cau cuu ho"\n-Background: do (khan cap)\n-Text color: white\n-Status: enable','','Untested'),
    ('GUI-GYC09','[Huy] Button huy form','Kiem tra nut huy','Da dang nhap','-Text: "Huy"\n-Background: xam\n-Status: enable','','Untested'),
    ['FUNCTION_SHOW Chuc nang gui yeu cau cuu ho'],
    ('FUNC-GYC01','Gui yeu cau thanh cong voi day du thong tin','1. Chon loai su co\n2. Nhap mo ta\n3. Lay vi tri GPS\n4. Chon muc khan cap\n5. Nhap so nguoi\n6. Click "Gui yeu cau"','Da dang nhap, la nguoi dan','Yeu cau duoc gui thanh cong, thong bao "Yeu cau da duoc tiep nhan", chuyen sang man hinh xem trang thai','','Untested'),
    ('FUNC-GYC02','Gui yeu cau khi chua chon loai su co','1. Bo trong Loai su co\n2. Dien cac truong con lai\n3. Click "Gui yeu cau"','Da dang nhap','Thong bao loi: "Vui long chon loai su co"','','Untested'),
    ('FUNC-GYC03','Gui yeu cau khi thieu vi tri','1. Chon loai su co\n2. Bo trong Vi tri\n3. Click "Gui yeu cau"','Da dang nhap','Thong bao loi: "Vui long cung cap vi tri de chung toi co the tim thay ban"','','Untested'),
    ('FUNC-GYC04','Gui yeu cau khi bo trong mo ta','1. Chon loai su co, vi tri\n2. Bo trong Mo ta\n3. Click "Gui yeu cau"','Da dang nhap','Thong bao loi: "Vui long mo ta tinh huong"','','Untested'),
    ('FUNC-GYC05','Lay vi tri GPS tu dong thanh cong','1. Click "Lay vi tri hien tai"\n2. Cap quyen truy cap vi tri khi duoc hoi','Da dang nhap, thiet bi co GPS','He thong tu dong dien toa do / dia chi hien tai vao field Vi tri','','Untested'),
    ('FUNC-GYC06','Lay vi tri GPS bi tu choi quyen','1. Click "Lay vi tri hien tai"\n2. Tu choi quyen vi tri','Da dang nhap, tu choi cap quyen GPS','Thong bao huong dan: "Vui long cap quyen truy cap vi tri de su dung tinh nang nay"','','Untested'),
    ('FUNC-GYC07','Gui yeu cau kem anh minh chung','1. Chon 3 anh minh chung\n2. Dien day du thong tin con lai\n3. Click "Gui yeu cau"','Da dang nhap','Yeu cau gui thanh cong kem theo anh, anh hien thi trong chi tiet yeu cau','','Untested'),
    ('FUNC-GYC08','Upload anh qua kich thuoc cho phep (>10MB)','1. Chon anh > 10MB\n2. Click "Gui yeu cau"','Da dang nhap','Thong bao loi: "Kich thuoc anh qua lon, vui long chon anh nho hon 10MB"','','Untested'),
    ('FUNC-GYC09','Gui yeu cau khi mat ket noi mang','1. Tat mang\n2. Dien day du thong tin\n3. Click "Gui yeu cau"','Da dang nhap, mat internet','Thong bao loi: "Khong co ket noi mang, vui long kiem tra va thu lai"','','Untested'),
    ('FUNC-GYC10','De trong tat ca cac truong - Click Gui','1. Khong nhap bat ky truong nao\n2. Click "Gui yeu cau"','Da dang nhap','He thong hien thi loi tren toan bo cac truong bat buoc','','Untested'),
    ('FUNC-GYC11','He thong AI tu dong tinh diem uu tien','1. Gui yeu cau khan cap voi vi tri cu the\n2. Kiem tra hang doi Admin','Yeu cau duoc tao thanh cong','Yeu cau xuat hien trong hang doi Admin voi diem uu tien duoc tinh tu dong boi AI','','Untested'),
    ('FUNC-GYC12','Nguoi dan nhan thong bao xac nhan qua email / push','1. Gui yeu cau thanh cong\n2. Kiem tra email / thong bao','Da gui yeu cau thanh cong','Nguoi dan nhan thong bao xac nhan kem ma yeu cau trong vong 1 phut','','Untested'),
])

# ========== SHEET: XEM TRANG THAI YEU CAU ==========
ws3 = wb.create_sheet('Xem Trang Thai Yeu Cau')
make_header(ws3, 'Xem Trang Thai Yeu Cau Cuu Ho')
write_rows(ws3, [
    ['GUI_SHOW Trang xem trang thai yeu cau'],
    ('GUI-XTTYC01','[Danh sach yeu cau] List / Card view','Kiem tra giao dien danh sach','Da dang nhap, co yeu cau truoc do','-Moi yeu cau hien thi: Ma yeu cau, Loai su co, Thoi gian gui, Trang thai\n-Sap xep moi nhat len tren\n-Status: enable','','Untested'),
    ('GUI-XTTYC02','[Badge trang thai] Mau sac phan loai trang thai','Kiem tra badge trang thai','Co yeu cau trong he thong','-"Cho xu ly": Mau vang\n-"Dang xu ly": Mau xanh duong\n-"Hoan thanh": Mau xanh la\n-"Da huy": Mau do\n-"Khong xu ly duoc": Mau xam','','Untested'),
    ('GUI-XTTYC03','[Chi tiet yeu cau] Trang chi tiet','Kiem tra chi tiet yeu cau','Co yeu cau trong he thong','-Hien thi: Ma yeu cau, Loai su co, Mo ta, Vi tri, Anh, Thoi gian, Trang thai\n-Thong tin doi cuu ho duoc phan cong (ten doi, so lien lac)\n-Ban do hien thi vi tri yeu cau','','Untested'),
    ('GUI-XTTYC04','[Nut Huy yeu cau] Chi hien thi khi con cho xu ly','Kiem tra hien thi nut huy','Yeu cau o trang thai Cho xu ly','-Nut "Huy yeu cau" hien thi va co the click\n-Trang thai khac: nut bi an hoac disable','','Untested'),
    ('GUI-XTTYC05','[Dong ho dem] Thoi gian cho xu ly','Kiem tra dong ho dem','Yeu cau o trang thai Cho xu ly','-Hien thi thoi gian da cho ke tu khi gui yeu cau\n-Cap nhat theo thoi gian thuc','','Untested'),
    ['FUNCTION_SHOW Chuc nang xem trang thai'],
    ('FUNC-XTTYC01','Xem danh sach yeu cau cua chinh minh','1. Dang nhap\n2. Vao muc "Yeu cau cua toi"','Da dang nhap, co yeu cau da gui','Chi hien thi cac yeu cau cua nguoi dung hien tai, khong hien thi yeu cau nguoi khac','','Untested'),
    ('FUNC-XTTYC02','Xem chi tiet mot yeu cau cu the','1. Click vao bat ky yeu cau nao trong danh sach','Co danh sach yeu cau','Mo trang chi tiet voi day du: ma yeu cau, loai, mo ta, vi tri tren ban do, trang thai, thong tin doi cuu ho','','Untested'),
    ('FUNC-XTTYC03','Trang thai cap nhat real-time khi duoc phan cong','1. Yeu cau dang "Cho xu ly"\n2. Admin phan cong doi cuu ho','Yeu cau dang o cho xu ly','Trang thai tu dong doi sang "Dang xu ly" va hien thi thong tin doi duoc phan cong (khong can reload)','','Untested'),
    ('FUNC-XTTYC04','Trang thai cap nhat real-time khi hoan thanh','1. Doi cuu ho nop bao cao hoan thanh\n2. Quan sat trang nguoi dan','Yeu cau dang xu ly','Trang thai tu dong doi sang "Hoan thanh" va hien thi form danh gia','','Untested'),
    ('FUNC-XTTYC05','Huy yeu cau khi chua duoc xu ly','1. Chon yeu cau o trang thai "Cho xu ly"\n2. Click "Huy yeu cau"\n3. Nhap ly do huy\n4. Xac nhan huy','Yeu cau o trang thai Cho xu ly','Yeu cau chuyen sang "Da huy", hien thi thong bao "Huy yeu cau thanh cong"','','Untested'),
    ('FUNC-XTTYC06','Khong the huy yeu cau khi dang xu ly','1. Click vao yeu cau o trang thai "Dang xu ly"\n2. Kiem tra tuy chon huy','Yeu cau dang xu ly','Nut "Huy yeu cau" bi an hoac vo hieu hoa, khong the huy','','Untested'),
    ('FUNC-XTTYC07','Loc yeu cau theo trang thai','1. Chon bo loc "Cho xu ly"\n2. Xem danh sach','Co yeu cau o nhieu trang thai khac nhau','Chi hien thi yeu cau co trang thai "Cho xu ly"','','Untested'),
    ('FUNC-XTTYC08','Loc yeu cau theo khoang thoi gian','1. Chon tu ngay - den ngay\n2. Ap dung bo loc','Co nhieu yeu cau trong he thong','Chi hien thi yeu cau trong khoang thoi gian da chon','','Untested'),
    ('FUNC-XTTYC09','Xem ban do vi tri yeu cau','1. Mo chi tiet yeu cau\n2. Xem tab "Vi tri"','Co yeu cau co thong tin GPS','Ban do hien thi dung vi tri da gui yeu cau, co the zoom va di chuyen ban do','','Untested'),
    ('FUNC-XTTYC10','Theo doi vi tri doi cuu ho tren ban do','1. Yeu cau dang o trang thai "Dang xu ly"\n2. Xem tab vi tri','Doi cuu ho dang di den hien truong','Ban do hien thi vi tri doi cuu ho dang di chuyen, cap nhat real-time','','Untested'),
])

# ========== SHEET: DANH GIA SAU CUU HO ==========
ws4 = wb.create_sheet('Danh Gia Sau Cuu Ho')
make_header(ws4, 'Danh Gia Dich Vu Sau Khi Cuu Ho')
write_rows(ws4, [
    ['GUI_SHOW Form danh gia dich vu cuu ho'],
    ('GUI-DG01','[Title] Tieu de form danh gia','Kiem tra giao dien form','Yeu cau da hoan thanh','-Title: "Danh gia dich vu cuu ho"\n-Mo ta: "Cam nhan cua ban giup chung toi cai thien dich vu"\n-Status: enable','','Untested'),
    ('GUI-DG02','[Danh gia sao] Star rating 1-5','Kiem tra UI sao danh gia','Yeu cau da hoan thanh','-5 ngoi sao co the click\n-Sao duoc chon hien thi mau vang\n-Hover hien thi mau nhat\n-Status: enable','','Untested'),
    ('GUI-DG03','[Nhan xet] Textbox','Kiem tra textbox nhan xet','Yeu cau da hoan thanh','-Placeholder: "Chia se cam nhan cua ban ve chat luong dich vu..."\n-Gioi han 300 ky tu\n-Hien thi so ky tu con lai\n-Status: enable','','Untested'),
    ('GUI-DG04','[Tag nhanh] Quick feedback tags','Kiem tra quick tags','Yeu cau da hoan thanh','-Cac tag nhanh: Phan hoi nhanh, Chuyen nghiep, Tan tinh, Giai quyet tot\n-Co the chon nhieu tag\n-Status: enable','','Untested'),
    ('GUI-DG05','[Gui danh gia] Button','Kiem tra nut gui danh gia','Yeu cau da hoan thanh','-Text: "Gui danh gia"\n-Background: xanh la\n-Vo hieu hoa khi chua chon sao\n-Status: enable','','Untested'),
    ('GUI-DG06','[Bo qua] Link bo qua danh gia','Kiem tra link bo qua','Yeu cau da hoan thanh','-Text: "Bo qua"\n-Cho phep nguoi dan bo qua danh gia\n-Status: enable','','Untested'),
    ['FUNCTION_SHOW Chuc nang danh gia sau cuu ho'],
    ('FUNC-DG01','Danh gia 5 sao kem nhan xet va tag','1. Chon 5 sao\n2. Nhap nhan xet\n3. Chon cac tag nhan xet nhanh\n4. Click "Gui danh gia"','Yeu cau da hoan thanh','Danh gia luu thanh cong, hien thi thong bao "Cam on ban da danh gia! Danh gia cua ban giup chung toi cai thien dich vu"','','Untested'),
    ('FUNC-DG02','Danh gia khi chua chon so sao','1. Khong chon sao\n2. Nhap nhan xet\n3. Click "Gui danh gia"','Yeu cau da hoan thanh','Nut "Gui danh gia" bi vo hieu hoa hoac thong bao loi: "Vui long chon so sao danh gia"','','Untested'),
    ('FUNC-DG03','Danh gia chi voi so sao - khong nhan xet','1. Chon 3 sao\n2. De trong o nhan xet\n3. Click "Gui danh gia"','Yeu cau da hoan thanh','Danh gia gui thanh cong (nhan xet la tuy chon)','','Untested'),
    ('FUNC-DG04','Danh gia 1 sao kem ly do','1. Chon 1 sao\n2. Nhap ly do khong hai long\n3. Click "Gui danh gia"','Yeu cau da hoan thanh','Danh gia luu thanh cong, Admin nhan canh bao danh gia thap','','Untested'),
    ('FUNC-DG05','Bo qua danh gia','1. Click link "Bo qua"','Yeu cau da hoan thanh, chua danh gia','He thong dong form danh gia, chuyen ve man hinh chinh, danh gia chua duoc luu','','Untested'),
    ('FUNC-DG06','Form danh gia khong hien thi khi yeu cau chua hoan thanh','1. Vao yeu cau dang o trang thai "Dang xu ly"\n2. Kiem tra form danh gia','Yeu cau chua hoan thanh','Form danh gia khong hien thi, thay vao do la trang thai hien tai cua yeu cau','','Untested'),
    ('FUNC-DG07','Khong the danh gia lan 2 cho cung mot yeu cau','1. Da danh gia yeu cau thanh cong\n2. Mo lai yeu cau do\n3. Kiem tra','Da danh gia lan 1','He thong hien thi danh gia da gui (so sao + nhan xet), khong co nut gui them','','Untested'),
    ('FUNC-DG08','Admin xem danh gia cua nguoi dan','1. Admin vao bao cao danh gia\n2. Xem danh gia chi tiet','Co danh gia tu nguoi dan','Admin thay danh gia voi: so sao, nhan xet, tag, thoi gian, thong tin yeu cau','','Untested'),
    ('FUNC-DG09','Thong ke danh gia trung binh cua moi doi cuu ho','1. Admin xem bao cao theo doi cuu ho','Co nhieu danh gia trong he thong','Hien thi diem danh gia trung binh cua moi doi, sap xep tu cao xuong thap','','Untested'),
])

wb.save(r'C:\xampp\htdocs\DATN\GR29\8.1.ProjectTestCaseSprint1_Complete.xlsx')
print('Sprint 1 Complete - OK')
print('Sheets:', wb.sheetnames)
